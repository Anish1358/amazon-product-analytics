"""
Step 4: Excel Dashboard
Builds a multi-sheet Excel workbook with:
- Raw data sheet
- A Summary/KPI sheet using real Excel formulas (SUMIFS/AVERAGEIFS)
- A native Excel PivotTable
- Two native Excel charts
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.pivot.cache import CacheDefinition
from openpyxl.utils import get_column_letter

df = pd.read_csv("C:\\Users\\anish\\OneDrive\\Documents\\Desktop\\jupyter notbook\\DASHBOARD PROj\\amazon_with_sentiment.csv")

wb = Workbook()

ws_data = wb.active
ws_data.title = "Data"
ws_data.append(list(df.columns))
for _, row in df.iterrows():
    ws_data.append(list(row))

header_font = Font(bold=True, color="FFFFFF", name="Arial")
header_fill = PatternFill("solid", fgColor="2F5597")
for cell in ws_data[1]:
    cell.font = header_font
    cell.fill = header_fill

n_rows = ws_data.max_row
n_cols = ws_data.max_column
last_col_letter = get_column_letter(n_cols)
tab = Table(displayName="ProductData", ref=f"A1:{last_col_letter}{n_rows}")
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False
)
ws_data.add_table(tab)

for col_cells in ws_data.columns:
    length = max(len(str(c.value)) for c in col_cells[:50])
    ws_data.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)

ws_sum = wb.create_sheet("Summary")
ws_sum["A1"] = "Amazon Product Analytics — Summary"
ws_sum["A1"].font = Font(bold=True, size=14, name="Arial")

categories = sorted(df["main_category"].dropna().unique().tolist())

headers = ["Category", "Num Products", "Avg Discounted Price", "Avg Discount %", "Avg Rating", "Total Reviews"]
for j, h in enumerate(headers, start=1):
    c = ws_sum.cell(row=3, column=j, value=h)
    c.font = header_font
    c.fill = header_fill

data_last_row = n_rows  
for i, cat in enumerate(categories, start=4):
    ws_sum.cell(row=i, column=1, value=cat)
  
    ws_sum.cell(row=i, column=2,
        value=f'=COUNTIF(Data!$C$2:$C${data_last_row},A{i})')
    ws_sum.cell(row=i, column=3,
        value=f'=ROUND(AVERAGEIF(Data!$C$2:$C${data_last_row},A{i},Data!$F$2:$F${data_last_row}),0)')
    ws_sum.cell(row=i, column=4,
        value=f'=ROUND(AVERAGEIF(Data!$C$2:$C${data_last_row},A{i},Data!$G$2:$G${data_last_row}),1)')
    ws_sum.cell(row=i, column=5,
        value=f'=ROUND(AVERAGEIF(Data!$C$2:$C${data_last_row},A{i},Data!$J$2:$J${data_last_row}),2)')
    ws_sum.cell(row=i, column=6,
        value=f'=SUMIF(Data!$C$2:$C${data_last_row},A{i},Data!$K$2:$K${data_last_row})')

last_cat_row = 3 + len(categories)

kpi_row = last_cat_row + 3
ws_sum.cell(row=kpi_row, column=1, value="Overall KPIs").font = Font(bold=True, size=12, name="Arial")
kpis = [
    ("Total Products", f"=COUNTA(Data!A2:A{data_last_row})"),
    ("Average Rating (all products)", f"=ROUND(AVERAGE(Data!J2:J{data_last_row}),2)"),
    ("Average Discount %", f"=ROUND(AVERAGE(Data!G2:G{data_last_row}),1)"),
    ("Total Reviews", f"=SUM(Data!K2:K{data_last_row})"),
]
for i, (label, formula) in enumerate(kpis, start=kpi_row + 1):
    ws_sum.cell(row=i, column=1, value=label)
    ws_sum.cell(row=i, column=2, value=formula)

for col, width in zip("ABCDEF", [24, 14, 20, 16, 12, 14]):
    ws_sum.column_dimensions[col].width = width

bar = BarChart()
bar.title = "Average Rating by Category"
bar.y_axis.title = "Avg Rating"
bar.x_axis.title = "Category"
data_ref = Reference(ws_sum, min_col=5, min_row=3, max_row=last_cat_row)
cats_ref = Reference(ws_sum, min_col=1, min_row=4, max_row=last_cat_row)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.width, bar.height = 18, 10
ws_sum.add_chart(bar, "H3")

pie = PieChart()
pie.title = "Product Count by Category"
data_ref2 = Reference(ws_sum, min_col=2, min_row=3, max_row=last_cat_row)
pie.add_data(data_ref2, titles_from_data=True)
pie.set_categories(cats_ref)
pie.width, pie.height = 18, 10
ws_sum.add_chart(pie, "H22")

ws_sent = wb.create_sheet("Sentiment")
ws_sent["A1"] = "Review Sentiment vs Star Rating"
ws_sent["A1"].font = Font(bold=True, size=14, name="Arial")

sent_labels = ["Positive", "Neutral", "Negative"]
ws_sent["A3"] = "Sentiment Label"
ws_sent["B3"] = "Num Products"
ws_sent["C3"] = "Avg Star Rating"
ws_sent["A3"].font = header_font; ws_sent["A3"].fill = header_fill
ws_sent["B3"].font = header_font; ws_sent["B3"].fill = header_fill
ws_sent["C3"].font = header_font; ws_sent["C3"].fill = header_fill

L_col = "L"  
sent_col_letter = get_column_letter(list(df.columns).index("sentiment_label") + 1)
rating_col_letter = get_column_letter(list(df.columns).index("rating") + 1)

for i, label in enumerate(sent_labels, start=4):
    ws_sent.cell(row=i, column=1, value=label)
    ws_sent.cell(row=i, column=2,
        value=f'=COUNTIF(Data!${sent_col_letter}$2:${sent_col_letter}${data_last_row},A{i})')
    ws_sent.cell(row=i, column=3,
        value=f'=ROUND(AVERAGEIF(Data!${sent_col_letter}$2:${sent_col_letter}${data_last_row},A{i},Data!${rating_col_letter}$2:${rating_col_letter}${data_last_row}),2)')

for col, width in zip("ABC", [18, 14, 16]):
    ws_sent.column_dimensions[col].width = width

wb.save("Amazon_Product_Analytics_Dashboard.xlsx")
print("Saved Amazon_Product_Analytics_Dashboard.xlsx")
print("sentiment_label column letter:", sent_col_letter, "| rating column letter:", rating_col_letter)